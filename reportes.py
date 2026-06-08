"""
reportes.py — Generación de archivos Excel y PDF con resultados BCRA.
Extraído de app.py para mantenerlo enfocado en UI.
"""
import io
import pandas as pd
from datetime import datetime
from bcra import calcular_pasa, detalle_str, periodo_a_texto

# ── Excel ─────────────────────────────────────────────────────────────────────

def escribir_hoja_con_formato(writer, sheet_name, titulo, df_data, totales=None):
    """Escribe una hoja Excel con título, encabezado, datos y fila de totales."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    ws = writer.book.create_sheet(title=sheet_name)
    writer.sheets[sheet_name] = ws

    ws.cell(row=1, column=1, value=titulo).font = Font(bold=True, size=12, color="FFFFFF")
    ws.cell(row=1, column=1).fill = PatternFill("solid", fgColor="0077B6")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df_data.columns))

    ws.cell(row=2, column=1,
            value=f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}").font = Font(italic=True, size=9, color="555555")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(df_data.columns))

    header_fill = PatternFill("solid", fgColor="00B4D8")
    for col_idx, col_name in enumerate(df_data.columns, 1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font      = Font(bold=True, color="FFFFFF", size=9)
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(df_data.itertuples(index=False), 4):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = Font(size=9)
            cell.alignment = Alignment(horizontal="left")
            if row_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="EAF4FB")
            col_name = df_data.columns[col_idx - 1]
            if col_name == "ESTADO_RIESGO":
                if value == "NO PASA":
                    cell.font = Font(bold=True, color="C62828", size=9)
                elif value == "PASA":
                    cell.font = Font(bold=True, color="006064", size=9)

    if totales is not None:
        total_row = len(df_data) + 4
        for col_idx, col_name in enumerate(df_data.columns, 1):
            cell = ws.cell(row=total_row, column=col_idx)
            if col_name in totales:
                cell.value = totales[col_name]
            elif col_idx == 1:
                cell.value = "TOTALES"
            cell.font = Font(bold=True, size=9, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="0077B6")

    for col_idx, col_name in enumerate(df_data.columns, 1):
        max_len = max(
            len(str(col_name)),
            max((len(str(v)) for v in df_data.iloc[:, col_idx - 1].fillna("")), default=0)
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[3].height = 16


def generar_excel(resultados: list, umbral: float) -> bytes:
    output = io.BytesIO()

    # Lista vacía — retornar Excel mínimo válido con hoja en blanco
    if not resultados:
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            pd.DataFrame({"Sin datos": []}).to_excel(writer, index=False, sheet_name="Sin datos")
        return output.getvalue()

    periodos     = [r.get("Periodo", "") for r in resultados if r.get("Periodo")]
    periodo_tit  = periodo_a_texto(max(set(periodos), key=periodos.count)) if periodos else ""
    fecha_consul = datetime.now().strftime('%d/%m/%Y %H:%M')

    with pd.ExcelWriter(output, engine="openpyxl") as writer:

        # ── Hoja 1: LOG por entidad ────────────────────────────────────────────
        rows_log = []
        for r in resultados:
            if r.get("error"):
                rows_log.append({
                    "CUIT": r["CUIT"], "DENOMINACION": r.get("Nombre", ""),
                    "PERIODO": "", "ENTIDAD": "ERROR API",
                    "SITUACION": "", "MONTO_BCRA_MILES": "",
                    "TIPO_SITUACION": "", "ESTADO_API": r["error"],
                })
                continue
            if not r.get("Entidades"):
                rows_log.append({
                    "CUIT": r["CUIT"], "DENOMINACION": r.get("Nombre", ""),
                    "PERIODO": r.get("Periodo", ""), "ENTIDAD": "SIN DEUDA REGISTRADA",
                    "SITUACION": "", "MONTO_BCRA_MILES": 0,
                    "TIPO_SITUACION": "SIN DEUDA", "ESTADO_API": "OK",
                })
                continue
            for ent in r["Entidades"]:
                sit  = ent["Situacion"]
                tipo = "NORMAL" if sit == 1 else ("RIESGO" if sit in [2, 3, 4, 5] else "OTRO")
                rows_log.append({
                    "CUIT": r["CUIT"], "DENOMINACION": r.get("Nombre", ""),
                    "PERIODO": r.get("Periodo", ""), "ENTIDAD": ent["Entidad"],
                    "SITUACION": sit, "MONTO_BCRA_MILES": ent["Monto"],
                    "TIPO_SITUACION": tipo, "ESTADO_API": "OK",
                })

        df_log  = pd.DataFrame(rows_log)
        if "MONTO_BCRA_MILES" in df_log.columns and not df_log.empty:
            tot_log = {"MONTO_BCRA_MILES": df_log["MONTO_BCRA_MILES"].apply(pd.to_numeric, errors="coerce").sum()}
        else:
            tot_log = {"MONTO_BCRA_MILES": 0}
        titulo_log = (
            f"LOG — Detalle por Entidad BCRA | Base BCRA: {periodo_tit} | "
            f"Consulta: {fecha_consul} | Umbral: {umbral}%"
        )
        escribir_hoja_con_formato(writer, "LOG Detalle", titulo_log, df_log, tot_log)

        # ── Hoja 2: Consolidado ────────────────────────────────────────────────
        rows2 = []
        for r in resultados:
            if r.get("error"):
                rows2.append({
                    "CUIT": r["CUIT"], "DENOMINACION": r.get("Nombre", ""),
                    "PERIODO": "", "CANT_OPERACIONES": r.get("Cant_Operaciones", 1),
                    "CANT_ENTIDADES": 0, "TOTAL_DEUDA_MILES": 0,
                    "TOTAL_SIT_NEGATIVAS_MILES": 0, "PORCENTAJE_NEGATIVO": "",
                    "ESTADO_RIESGO": "ERROR", "DETALLE_DEUDA": r["error"],
                })
                continue
            ratio, resultado = calcular_pasa(r["Monto_Sit1"], r["Monto_Riesgo"], umbral)
            total_deuda = r["Monto_Sit1"] + r["Monto_Riesgo"]
            rows2.append({
                "CUIT":                      r["CUIT"],
                "DENOMINACION":              r.get("Nombre", ""),
                "PERIODO":                   r.get("Periodo", ""),
                "CANT_OPERACIONES":          r.get("Cant_Operaciones", 1),
                "CANT_ENTIDADES":            len(r.get("Entidades", [])),
                "TOTAL_DEUDA_MILES":         total_deuda,
                "TOTAL_SIT_NEGATIVAS_MILES": r["Monto_Riesgo"],
                "PORCENTAJE_NEGATIVO":       f"{ratio:.2f}%",
                "ESTADO_RIESGO":             resultado,
                "DETALLE_DEUDA":             detalle_str(r.get("Entidades", [])),
            })

        df2      = pd.DataFrame(rows2)
        pasan    = (df2["ESTADO_RIESGO"] == "PASA").sum()
        no_pasan = (df2["ESTADO_RIESGO"] == "NO PASA").sum()
        errores  = (df2["ESTADO_RIESGO"] == "ERROR").sum()
        titulo2 = (
            f"Consolidado BCRA | Base BCRA: {periodo_tit} | Consulta: {fecha_consul} | "
            f"Umbral: {umbral}% | Total: {len(df2)} | Pasan: {pasan} | "
            f"No pasan: {no_pasan} | Errores: {errores}"
        )
        totales2 = {
            "CANT_OPERACIONES":          df2["CANT_OPERACIONES"].sum() if not df2.empty else 0,
            "CANT_ENTIDADES":            df2["CANT_ENTIDADES"].sum() if not df2.empty else 0,
            "TOTAL_DEUDA_MILES":         df2["TOTAL_DEUDA_MILES"].apply(pd.to_numeric, errors="coerce").sum() if not df2.empty else 0,
            "TOTAL_SIT_NEGATIVAS_MILES": df2["TOTAL_SIT_NEGATIVAS_MILES"].apply(pd.to_numeric, errors="coerce").sum() if not df2.empty else 0,
        }
        escribir_hoja_con_formato(writer, "Consolidado Actual", titulo2, df2, totales2)

        # ── Hoja 3: Con Capital Vendido (opcional) ─────────────────────────────
        rows3 = [r for r in resultados if not r.get("error") and r.get("Capital") is not None]
        if rows3:
            data3 = []
            for r in rows3:
                sit1_total = r["Monto_Sit1"] + r["Capital"]
                total      = sit1_total + r["Monto_Riesgo"]
                ratio_cc   = round(r["Monto_Riesgo"] / total * 100, 2) if total > 0 else 0
                res_cc     = "NO PASA" if ratio_cc >= umbral else "PASA"
                data3.append({
                    "CUIT":                      r["CUIT"],
                    "DENOMINACION":              r.get("Nombre", ""),
                    "PERIODO":                   r.get("Periodo", ""),
                    "CANT_OPERACIONES":          r.get("Cant_Operaciones", 1),
                    "CANT_ENTIDADES":            len(r.get("Entidades", [])),
                    "CAPITAL_VENDIDO":           r["Capital"],
                    "TOTAL_DEUDA_BCRA_MILES":    r["Monto_Sit1"] + r["Monto_Riesgo"],
                    "SIT1_BCRA_MILES":           r["Monto_Sit1"],
                    "SIT1_TOTAL_MILES":          sit1_total,
                    "TOTAL_SIT_NEGATIVAS_MILES": r["Monto_Riesgo"],
                    "PORCENTAJE_NEGATIVO":       f"{ratio_cc:.2f}%",
                    "ESTADO_RIESGO":             res_cc,
                    "DETALLE_DEUDA":             detalle_str(r.get("Entidades", [])),
                })
            df3     = pd.DataFrame(data3)
            titulo3 = f"Con Capital Vendido | Base BCRA: {periodo_tit} | Consulta: {fecha_consul} | Umbral: {umbral}%"
            totales3 = {
                "CAPITAL_VENDIDO":           df3["CAPITAL_VENDIDO"].sum(),
                "TOTAL_DEUDA_BCRA_MILES":    df3["TOTAL_DEUDA_BCRA_MILES"].sum(),
                "TOTAL_SIT_NEGATIVAS_MILES": df3["TOTAL_SIT_NEGATIVAS_MILES"].sum(),
            }
            escribir_hoja_con_formato(writer, "Con Capital Vendido", titulo3, df3, totales3)

    return output.getvalue()

# ── PDF ───────────────────────────────────────────────────────────────────────

def generar_pdf(resultados: list, umbral: float, graf_torta: dict = None, graf_barras: dict = None, empresa: dict = None) -> bytes | None:
    """Genera un PDF personalizado con logo y datos de empresa + gráficos + 3 secciones de detalle."""
    try:
        from reportlab.lib.pagesizes import landscape, A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.lib import colors
        from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                        Table, TableStyle, PageBreak)
    except ImportError:
        return None

    AVISO_LEGAL = (
        "Aviso: Los datos provienen de fuentes publicas (BCRA y boletines oficiales). "
        "Nuestro sistema agrega valor mediante procesamiento y presentacion, pero no modifica "
        "la informacion. El BCRA no avala ni certifica este servicio."
    )

    def _encabezado_pagina(canvas, doc):
        """
        Encabezado de cada página:
          - Nombre empresa en negrita, arriba izquierda
          - Logo arriba derecha
        Pie de cada página:
          - Izquierda: domicilio · ciudad · provincia
          - Centro:    teléfono · email · web
          - Derecha:   aviso legal + pág
        """
        canvas.saveState()
        page_w, page_h = landscape(A4)

        # ── ENCABEZADO ───────────────────────────────────────────────────────
        y_header = page_h - 0.9*cm

        if empresa:
            nombre_emp = empresa.get("nombre","")
            cuit_emp   = f"  ·  CUIT {empresa['cuit_empresa']}" if empresa.get("cuit_empresa") else ""

            # Nombre empresa — arriba izquierda, negrita
            canvas.setFont("Helvetica-Bold", 9)
            canvas.setFillColor(colors.HexColor("#1d1d1f"))
            canvas.drawString(1.2*cm, y_header, f"{nombre_emp}{cuit_emp}")

        # Logo — arriba derecha
        if empresa and empresa.get("logo_bytes"):
            try:
                from reportlab.lib.utils import ImageReader
                img_buf = io.BytesIO(empresa["logo_bytes"])
                img_rdr = ImageReader(img_buf)
                iw, ih  = img_rdr.getSize()
                max_h   = 1.0*cm
                ratio   = max_h / ih
                logo_w  = iw * ratio
                canvas.drawImage(img_rdr,
                                 page_w - 1.2*cm - logo_w, y_header - 0.1*cm,
                                 width=logo_w, height=max_h,
                                 preserveAspectRatio=True, mask="auto")
            except Exception:
                pass

        # Línea bajo encabezado
        canvas.setStrokeColor(colors.HexColor("#e5e5ea"))
        canvas.setLineWidth(0.4)
        canvas.line(1.2*cm, y_header - 0.2*cm, page_w - 1.2*cm, y_header - 0.2*cm)

        # ── PIE ──────────────────────────────────────────────────────────────
        y_line = 1.4*cm
        y_pie1 = 1.0*cm
        y_pie2 = 0.55*cm

        canvas.setStrokeColor(colors.HexColor("#d2d2d7"))
        canvas.setLineWidth(0.4)
        canvas.line(1.2*cm, y_line, page_w - 1.2*cm, y_line)

        if empresa:
            dom_parts  = filter(None, [empresa.get("domicilio"), empresa.get("ciudad"), empresa.get("provincia")])
            dom_str    = " · ".join(dom_parts)
            cont_parts = filter(None, [empresa.get("telefono"), empresa.get("email_empresa"), empresa.get("web")])
            cont_str   = " · ".join(cont_parts)

            canvas.setFont("Helvetica", 6)
            canvas.setFillColor(colors.HexColor("#86868b"))
            if dom_str:
                canvas.drawString(1.2*cm, y_pie1, dom_str)
            if cont_str:
                canvas.drawString(1.2*cm, y_pie2, cont_str)

        # Aviso legal centrado
        canvas.setFont("Helvetica", 5.5)
        canvas.setFillColor(colors.HexColor("#aeaeb2"))
        canvas.drawCentredString(page_w / 2, y_pie1, "Datos de fuentes publicas (BCRA y boletines oficiales).")
        canvas.drawCentredString(page_w / 2, y_pie2, "El BCRA no avala ni certifica este servicio. Ley 25.326.")

        # Número de página — derecha
        canvas.setFont("Helvetica-Bold", 6.5)
        canvas.setFillColor(colors.HexColor("#0066cc"))
        canvas.drawRightString(page_w - 1.2*cm, y_pie1, f"Deudix")
        canvas.drawRightString(page_w - 1.2*cm, y_pie2, f"Pág. {doc.page}")

        canvas.restoreState()

    buffer  = io.BytesIO()
    doc     = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                                rightMargin=1.2*cm, leftMargin=1.2*cm,
                                topMargin=1.8*cm,  bottomMargin=2.0*cm)
    styles  = getSampleStyleSheet()
    C_AZUL  = colors.HexColor("#0077b6")
    C_CYAN  = colors.HexColor("#00b4d8")
    C_ROJO  = colors.HexColor("#c62828")
    C_VERDE = colors.HexColor("#006064")
    C_GRIS  = colors.HexColor("#e8f4f8")
    C_BLANC = colors.white

    title_style = ParagraphStyle("TIT", parent=styles["Title"], fontSize=13,
                                 textColor=C_CYAN, spaceAfter=2, spaceBefore=0)
    sub_style   = ParagraphStyle("SUB", parent=styles["Normal"], fontSize=8,
                                 textColor=colors.HexColor("#555555"), spaceAfter=4)
    sec_style   = ParagraphStyle("SEC", parent=styles["Heading2"], fontSize=10,
                                 textColor=C_AZUL, spaceBefore=10, spaceAfter=4)

    periodos   = [r.get("Periodo", "") for r in resultados if r.get("Periodo")]
    periodo_txt = periodo_a_texto(max(set(periodos), key=periodos.count)) if periodos else ""
    pasan    = sum(1 for r in resultados if not r.get("error")
                   and calcular_pasa(r["Monto_Sit1"], r["Monto_Riesgo"], umbral)[1] == "PASA")
    no_pasan = sum(1 for r in resultados if not r.get("error")
                   and calcular_pasa(r["Monto_Sit1"], r["Monto_Riesgo"], umbral)[1] == "NO PASA")
    errores  = sum(1 for r in resultados if r.get("error"))

    def tabla_estilo(data, col_widths, col_estado=None):
        t = Table(data, colWidths=col_widths, repeatRows=1)
        estilo = [
            ("BACKGROUND",    (0, 0), (-1, 0),  C_AZUL),
            ("TEXTCOLOR",     (0, 0), (-1, 0),  C_BLANC),
            ("FONTNAME",      (0, 0), (-1, 0),  "Helvetica-Bold"),
            ("FONTSIZE",      (0, 0), (-1, -1), 6.5),
            ("ROWBACKGROUNDS",(0, 1), (-1, -2), [C_BLANC, C_GRIS]),
            ("BACKGROUND",    (0,-1), (-1, -1), C_AZUL),
            ("TEXTCOLOR",     (0,-1), (-1, -1), C_BLANC),
            ("FONTNAME",      (0,-1), (-1, -1), "Helvetica-Bold"),
            ("GRID",          (0, 0), (-1, -1), 0.3, colors.HexColor("#90caf9")),
            ("ALIGN",         (0, 0), (-1, -1), "LEFT"),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING",    (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ]
        t.setStyle(TableStyle(estilo))
        if col_estado is not None:
            for i, row in enumerate(data[1:-1], 1):
                val = row[col_estado]
                if val == "NO PASA":
                    t.setStyle(TableStyle([("TEXTCOLOR",  (col_estado, i), (col_estado, i), C_ROJO),
                                           ("FONTNAME",   (col_estado, i), (col_estado, i), "Helvetica-Bold")]))
                elif val == "PASA":
                    t.setStyle(TableStyle([("TEXTCOLOR",  (col_estado, i), (col_estado, i), C_VERDE),
                                           ("FONTNAME",   (col_estado, i), (col_estado, i), "Helvetica-Bold")]))
                elif val == "RIESGO":
                    t.setStyle(TableStyle([("TEXTCOLOR",  (col_estado, i), (col_estado, i), C_ROJO)]))
        return t

    # ── PÁGINA 0: Gráficos de resumen ────────────────────────────────────────
    # Intentamos generar imágenes con matplotlib (opcional).
    # Si no está disponible, los gráficos se reemplazan por tablas de texto.
    def _imagen_torta(labels, valores, colores, ancho_cm, alto_cm):
        """Retorna un objeto Image de reportlab o None."""
        try:
            import matplotlib
            matplotlib.use("Agg")
            import matplotlib.pyplot as plt
            import matplotlib.patches as mpatches
            from reportlab.platypus import Image as RLImage
            fig, ax = plt.subplots(figsize=(ancho_cm / 2.54, alto_cm / 2.54))
            wedges, texts, autotexts = ax.pie(
                valores, labels=labels, colors=colores,
                autopct="%1.1f%%", startangle=90,
                wedgeprops=dict(width=0.55),
                textprops=dict(fontsize=9),
            )
            for at in autotexts:
                at.set_fontsize(8)
            ax.set_title("Distribución de resultados", fontsize=10, pad=10)
            buf = io.BytesIO()
            fig.savefig(buf, format="png", bbox_inches="tight", dpi=150, transparent=True)
            plt.close(fig)
            buf.seek(0)
            return RLImage(buf, width=ancho_cm * cm, height=alto_cm * cm)
        except Exception:
            return None

    def _imagen_barras(sit_conteo, ancho_cm, alto_cm):
        try:
            import matplotlib
            matplotlib.use("Agg")
            import matplotlib.pyplot as plt
            from reportlab.platypus import Image as RLImage
            sit_labels  = ["Sit.1 Normal", "Sit.2", "Sit.3", "Sit.4", "Sit.5 Irrecup."]
            sit_colores = ["#1a7a1a", "#e68a00", "#cc5500", "#cc0000", "#7a0000"]
            valores = [sit_conteo.get(s, 0) for s in range(1, 6)]
            fig, ax = plt.subplots(figsize=(ancho_cm / 2.54, alto_cm / 2.54))
            bars = ax.bar(sit_labels, valores, color=sit_colores, width=0.55)
            for bar, v in zip(bars, valores):
                if v > 0:
                    ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.3,
                            str(v), ha="center", va="bottom", fontsize=8)
            ax.set_title("Entidades por situación BCRA", fontsize=10, pad=10)
            ax.set_ylabel("Cantidad de entidades", fontsize=8)
            ax.tick_params(axis="both", labelsize=8)
            ax.spines[["top", "right"]].set_visible(False)
            ax.set_facecolor("#fafafa")
            buf = io.BytesIO()
            fig.savefig(buf, format="png", bbox_inches="tight", dpi=150, transparent=True)
            plt.close(fig)
            buf.seek(0)
            return RLImage(buf, width=ancho_cm * cm, height=alto_cm * cm)
        except Exception:
            return None

    story = []
    nombre_empresa_pdf = (empresa.get("nombre","") if empresa else "") or ""
    titulo_pdf = f"{nombre_empresa_pdf} · Consulta BCRA" if nombre_empresa_pdf else "Sistema de Consulta BCRA"
    story.append(Paragraph(titulo_pdf, title_style))
    story.append(Paragraph(
        f"Consulta realizada: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  "
        f"Base BCRA activa: {periodo_txt}  |  Umbral: {umbral}%",
        sub_style))
    story.append(Spacer(1, 0.3*cm))

    # Gráficos de resumen — ocupan toda la primera página
    sin_deuda_count = sum(1 for r in resultados if not r.get("error") and r.get("Sin_Deuda"))
    img_torta  = None
    img_barras = None

    if graf_torta and graf_torta.get("valores"):
        # Usar los colores exactos que se mostraron en pantalla
        colores_pasados = graf_torta.get("colores",
            ["#1a7a1a","#cc0000","#0066cc","#86868b"][:len(graf_torta["labels"])])
        img_torta = _imagen_torta(
            graf_torta["labels"], graf_torta["valores"],
            colores_pasados, 12, 9
        )
    if graf_barras:
        img_barras = _imagen_barras(graf_barras, 12, 9)

    if img_torta or img_barras:
        from reportlab.platypus import HRFlowable
        story.append(Paragraph("Resumen visual del procesamiento", sec_style))
        story.append(Spacer(1, 0.2*cm))

        # Tabla de 2 columnas con los dos gráficos
        graf_cells = [img_torta or Paragraph("(gráfico no disponible)", sub_style),
                      img_barras or Paragraph("(gráfico no disponible)", sub_style)]
        t_graf = Table([[graf_cells[0], graf_cells[1]]], colWidths=[13*cm, 13*cm])
        t_graf.setStyle(TableStyle([
            ("ALIGN",   (0,0), (-1,-1), "CENTER"),
            ("VALIGN",  (0,0), (-1,-1), "MIDDLE"),
            ("LEFTPADDING",  (0,0), (-1,-1), 8),
            ("RIGHTPADDING", (0,0), (-1,-1), 8),
        ]))
        story.append(t_graf)
        story.append(Spacer(1, 0.4*cm))
        story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#90caf9")))
        story.append(Spacer(1, 0.2*cm))

    # Resumen estadístico en texto
    resumen_data = [
        ["TOTAL", "PASAN", "NO PASAN", "SIN DEUDA", "ERRORES"],
        [str(len(resultados)), str(pasan), str(no_pasan),
         str(sin_deuda_count),
         str(errores)]
    ]
    t_res = Table(resumen_data, colWidths=[3.5*cm]*5)
    t_res.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0), colors.HexColor("#0a0e1a")),
        ("TEXTCOLOR",     (0, 0), (-1, 0), C_CYAN),
        ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, 0), 7),
        ("ALIGN",         (0, 0), (-1, 0), "CENTER"),
        ("BACKGROUND",    (0, 1), (-1, 1), colors.HexColor("#111827")),
        ("FONTNAME",      (0, 1), (-1, 1), "Helvetica-Bold"),
        ("FONTSIZE",      (0, 1), (-1, 1), 18),
        ("ALIGN",         (0, 1), (-1, 1), "CENTER"),
        ("VALIGN",        (0, 0), (-1, -1),"MIDDLE"),
        ("TOPPADDING",    (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("TEXTCOLOR",     (0, 1), (0, 1),  colors.HexColor("#00b4d8")),
        ("TEXTCOLOR",     (1, 1), (1, 1),  colors.HexColor("#00e5ff")),
        ("TEXTCOLOR",     (2, 1), (2, 1),  colors.HexColor("#ff6b6b")),
        ("TEXTCOLOR",     (3, 1), (3, 1),  colors.HexColor("#ffd166")),
        ("TEXTCOLOR",     (4, 1), (4, 1),  colors.HexColor("#7eb8d4")),
        ("BOX",           (0, 0), (-1, -1),1, colors.HexColor("#1e3a5f")),
        ("INNERGRID",     (0, 0), (-1, -1),0.5, colors.HexColor("#1e3a5f")),
    ]))
    story.append(t_res)
    story.append(Spacer(1, 0.3*cm))

    # Sección 1 — LOG
    story.append(Paragraph(f"1. LOG — Detalle por Entidad  |  Período: {periodo_txt}", sec_style))
    log_rows   = [["CUIT","DENOMINACION","PERIODO","ENTIDAD","SIT.","MONTO $","TIPO"]]
    tot_monto  = 0.0
    for r in resultados:
        if r.get("error"):
            log_rows.append([str(r["CUIT"]), str(r.get("Nombre",""))[:22], "", "ERROR API", "", "", ""])
            continue
        if not r.get("Entidades"):
            log_rows.append([str(r["CUIT"]), str(r.get("Nombre",""))[:22],
                             str(r.get("Periodo","")), "SIN DEUDA", "", "$0", "SIN DEUDA"])
            continue
        for ent in r["Entidades"]:
            sit   = ent["Situacion"]
            tipo  = "NORMAL" if sit == 1 else ("RIESGO" if sit in [2,3,4,5] else "OTRO")
            monto = float(ent["Monto"])
            tot_monto += monto
            log_rows.append([
                str(r["CUIT"]), str(r.get("Nombre",""))[:22], str(r.get("Periodo","")),
                str(ent["Entidad"])[:35], str(sit), f"${monto:,.0f}", tipo,
            ])
    log_rows.append([
        "TOTALES", f"{len(resultados)} CUITs", "",
        f"{sum(len(r.get('Entidades',[])) for r in resultados if not r.get('error'))} entidades",
        "", f"${tot_monto:,.0f}", "",
    ])
    story.append(tabla_estilo(log_rows, [2.5*cm, 4.5*cm, 1.8*cm, 7*cm, 1.2*cm, 2.5*cm, 2*cm], col_estado=6))
    story.append(PageBreak())

    # Sección 2 — Consolidado
    story.append(Paragraph(f"2. Consolidado Actual  |  Período: {periodo_txt}  |  Umbral: {umbral}%", sec_style))
    cons_rows  = [["CUIT","DENOMINACION","PERÍODO","CANT.ENT.","DEUDA TOTAL $","SIT.NEG. $","% NEG.","ESTADO","DETALLE"]]
    tot_deuda  = tot_neg = tot_ent = 0.0
    for r in resultados:
        if r.get("error"):
            cons_rows.append([str(r["CUIT"]), r.get("Nombre","")[:18], "", "0",
                               "$0","$0","","ERROR",""])
            continue
        ratio, resultado = calcular_pasa(r["Monto_Sit1"], r["Monto_Riesgo"], umbral)
        td  = r["Monto_Sit1"] + r["Monto_Riesgo"]
        ce  = len(r.get("Entidades", []))
        tot_deuda += td; tot_neg += r["Monto_Riesgo"]; tot_ent += ce
        det = detalle_str(r.get("Entidades", []))
        det = det[:70] + "..." if len(det) > 70 else det
        cons_rows.append([
            str(r["CUIT"]), str(r.get("Nombre",""))[:18], str(r.get("Periodo","")),
            str(ce), f"${td:,.0f}", f"${r['Monto_Riesgo']:,.0f}",
            f"{ratio:.1f}%", resultado, det,
        ])
    cons_rows.append(["TOTALES", f"Pasan: {pasan}  No pasan: {no_pasan}  Err: {errores}",
                      "", str(int(tot_ent)), f"${tot_deuda:,.0f}", f"${tot_neg:,.0f}", "", "", ""])
    story.append(tabla_estilo(cons_rows,
        [2.5*cm,3.8*cm,1.6*cm,1.4*cm,2.8*cm,2.8*cm,1.6*cm,2*cm,7*cm], col_estado=7))

    # Sección 3 — Con Capital Vendido
    rows_cc = [r for r in resultados if not r.get("error") and r.get("Capital") is not None]
    if rows_cc:
        story.append(PageBreak())
        story.append(Paragraph(f"3. Con Capital Vendido  |  Período: {periodo_txt}  |  Umbral: {umbral}%", sec_style))
        cap_rows  = [["CUIT","DENOMINACION","PERÍODO","CAPITAL VENDIDO","SIT1 BCRA",
                      "SIT1 TOTAL","SIT.NEG.","% NEG.","ESTADO"]]
        tot_cap = tot_s1 = tot_s1t = tot_n3 = 0.0
        for r in rows_cc:
            sit1_total = r["Monto_Sit1"] + r["Capital"]
            total      = sit1_total + r["Monto_Riesgo"]
            ratio_cc   = round(r["Monto_Riesgo"] / total * 100, 2) if total > 0 else 0
            res_cc     = "NO PASA" if ratio_cc >= umbral else "PASA"
            tot_cap   += r["Capital"]; tot_s1  += r["Monto_Sit1"]
            tot_s1t   += sit1_total;   tot_n3  += r["Monto_Riesgo"]
            cap_rows.append([
                str(r["CUIT"]), str(r.get("Nombre",""))[:18], str(r.get("Periodo","")),
                f"${r['Capital']:,.0f}", f"${r['Monto_Sit1']:,.0f}",
                f"${sit1_total:,.0f}", f"${r['Monto_Riesgo']:,.0f}",
                f"{ratio_cc:.1f}%", res_cc,
            ])
        cap_rows.append(["TOTALES","","", f"${tot_cap:,.0f}", f"${tot_s1:,.0f}",
                         f"${tot_s1t:,.0f}", f"${tot_n3:,.0f}", "", ""])
        story.append(tabla_estilo(cap_rows,
            [2.5*cm,3.8*cm,1.6*cm,3*cm,2.8*cm,2.8*cm,2.8*cm,1.8*cm,2*cm], col_estado=8))

    doc.build(story, onFirstPage=_encabezado_pagina, onLaterPages=_encabezado_pagina)
    return buffer.getvalue()


# ── PDF de Seguimiento Mensual ─────────────────────────────────────────────────

def generar_pdf_seguimiento(resultados: list, empresa: dict, periodo: str) -> bytes | None:
    """
    Genera el PDF mensual del seguimiento de CUITs vigilados.
    Una página por cliente con resumen de variaciones.
    """
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.lib import colors
        from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                        Table, TableStyle, PageBreak)
    except ImportError:
        return None

    buffer    = io.BytesIO()
    C_AZUL    = colors.HexColor("#0066cc")
    C_VERDE   = colors.HexColor("#1a7a1a")
    C_ROJO    = colors.HexColor("#cc0000")
    C_GRIS    = colors.HexColor("#86868b")
    C_FONDO   = colors.HexColor("#f5f5f7")
    C_BLANC   = colors.white
    C_OSCURO  = colors.HexColor("#1d1d1f")

    AVISO = (
        "Datos de fuentes publicas (BCRA). "
        "El BCRA no avala ni certifica este servicio. Ley 25.326."
    )

    nombre_emp = (empresa or {}).get("nombre", "") if empresa else ""
    cuit_emp   = (empresa or {}).get("cuit_empresa", "") if empresa else ""

    def _pie(canvas, doc):
        canvas.saveState()
        pw, ph = A4
        canvas.setStrokeColor(colors.HexColor("#d2d2d7"))
        canvas.setLineWidth(0.4)
        canvas.line(1.5*cm, 1.2*cm, pw - 1.5*cm, 1.2*cm)
        canvas.setFont("Helvetica-Bold", 7)
        canvas.setFillColor(C_OSCURO)
        canvas.drawString(1.5*cm, 0.85*cm, nombre_emp)
        canvas.setFont("Helvetica", 6)
        canvas.setFillColor(C_GRIS)
        canvas.drawCentredString(pw/2, 0.85*cm, AVISO)
        canvas.setFont("Helvetica-Bold", 6)
        canvas.setFillColor(C_AZUL)
        canvas.drawRightString(pw - 1.5*cm, 0.85*cm, f"Pág. {doc.page}")
        canvas.restoreState()

    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        rightMargin=1.5*cm, leftMargin=1.5*cm,
        topMargin=1.5*cm, bottomMargin=1.8*cm,
    )
    styles = getSampleStyleSheet()

    title_s  = ParagraphStyle("T",  parent=styles["Title"],   fontSize=16,
                               textColor=C_OSCURO, spaceAfter=4, spaceBefore=0)
    sub_s    = ParagraphStyle("S",  parent=styles["Normal"],  fontSize=9,
                               textColor=C_GRIS,   spaceAfter=12)
    sec_s    = ParagraphStyle("SE", parent=styles["Heading2"],fontSize=11,
                               textColor=C_AZUL,   spaceBefore=14, spaceAfter=6)

    story = []

    # Encabezado
    story.append(Paragraph(f"Informe de Seguimiento Mensual", title_s))
    mes_txt = periodo_a_texto(periodo.replace("-","")) if len(periodo) == 7 else periodo
    story.append(Paragraph(
        f"{nombre_emp}{(' · CUIT ' + cuit_emp) if cuit_emp else ''} &nbsp;&nbsp;|&nbsp;&nbsp; "
        f"Período: {mes_txt} &nbsp;&nbsp;|&nbsp;&nbsp; "
        f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        sub_s,
    ))

    # Resumen estadístico
    suben      = sum(1 for r in resultados if r.get("variacion") == "SUBE")
    bajan      = sum(1 for r in resultados if r.get("variacion") == "BAJA")
    sin_cambio = sum(1 for r in resultados if r.get("variacion") == "SIN_CAMBIO")
    nuevos     = sum(1 for r in resultados if r.get("variacion") == "NUEVO")
    errores    = sum(1 for r in resultados if r.get("variacion") == "ERROR")

    res_data = [
        ["TOTAL", "SUBEN", "BAJAN", "SIN CAMBIO", "NUEVOS", "ERRORES"],
        [str(len(resultados)), str(suben), str(bajan),
         str(sin_cambio), str(nuevos), str(errores)],
    ]
    t_res = Table(res_data, colWidths=[2.8*cm]*6)
    t_res.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,0),  C_OSCURO),
        ("TEXTCOLOR",     (0,0), (-1,0),  C_AZUL),
        ("FONTNAME",      (0,0), (-1,0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0,0), (-1,0),  7),
        ("ALIGN",         (0,0), (-1,-1), "CENTER"),
        ("BACKGROUND",    (0,1), (-1,1),  C_FONDO),
        ("FONTNAME",      (0,1), (-1,1),  "Helvetica-Bold"),
        ("FONTSIZE",      (0,1), (-1,1),  16),
        ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING",    (0,0), (-1,-1), 6),
        ("BOTTOMPADDING", (0,0), (-1,-1), 6),
        ("TEXTCOLOR",     (0,1), (0,1),   C_OSCURO),
        ("TEXTCOLOR",     (1,1), (1,1),   C_ROJO),
        ("TEXTCOLOR",     (2,1), (2,1),   C_VERDE),
        ("BOX",           (0,0), (-1,-1), 0.5, colors.HexColor("#e5e5ea")),
        ("INNERGRID",     (0,0), (-1,-1), 0.3, colors.HexColor("#e5e5ea")),
    ]))
    story.append(t_res)
    story.append(Spacer(1, 0.4*cm))

    def _tabla_variacion(titulo, items, color_var):
        if not items:
            return
        story.append(Paragraph(titulo, sec_s))
        filas  = [["CUIT", "Nombre / Alias", "Deuda Normal $",
                    "Deuda Riesgo $", "Período", "Variación"]]
        for r in items:
            res      = r.get("resultado") or {}
            variacion = r.get("variacion","")
            filas.append([
                str(r.get("cuit","")),
                str(r.get("alias", r.get("cuit","")))[:35],
                f"${res.get('Monto_Sit1',0):,.0f}",
                f"${res.get('Monto_Riesgo',0):,.0f}",
                str(res.get("Periodo","")),
                VAR_LABEL.get(variacion, variacion),
            ])
        t = Table(filas, colWidths=[2.8*cm, 5.5*cm, 2.5*cm, 2.5*cm, 1.8*cm, 2.2*cm])
        estilo = [
            ("BACKGROUND",    (0,0), (-1,0),  C_AZUL),
            ("TEXTCOLOR",     (0,0), (-1,0),  C_BLANC),
            ("FONTNAME",      (0,0), (-1,0),  "Helvetica-Bold"),
            ("FONTSIZE",      (0,0), (-1,-1), 7),
            ("ROWBACKGROUNDS",(0,1), (-1,-1), [C_BLANC, C_FONDO]),
            ("GRID",          (0,0), (-1,-1), 0.3, colors.HexColor("#e5e5ea")),
            ("ALIGN",         (2,0), (4,-1),  "RIGHT"),
            ("ALIGN",         (0,0), (1,-1),  "LEFT"),
            ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
            ("TOPPADDING",    (0,0), (-1,-1), 3),
            ("BOTTOMPADDING", (0,0), (-1,-1), 3),
            ("TEXTCOLOR",     (-1,1),(-1,-1), color_var),
            ("FONTNAME",      (-1,1),(-1,-1), "Helvetica-Bold"),
        ]
        t.setStyle(TableStyle(estilo))
        story.append(t)
        story.append(Spacer(1, 0.2*cm))

    # Secciones por tipo de variación
    _tabla_variacion(
        f"CUITs con deuda que subió ({suben})",
        [r for r in resultados if r.get("variacion") == "SUBE"],
        C_ROJO,
    )
    _tabla_variacion(
        f"CUITs con deuda que bajó ({bajan})",
        [r for r in resultados if r.get("variacion") == "BAJA"],
        C_VERDE,
    )
    _tabla_variacion(
        f"CUITs sin cambio ({sin_cambio})",
        [r for r in resultados if r.get("variacion") == "SIN_CAMBIO"],
        C_GRIS,
    )
    _tabla_variacion(
        f"CUITs nuevos en seguimiento ({nuevos})",
        [r for r in resultados if r.get("variacion") == "NUEVO"],
        C_AZUL,
    )

    errores_list = [r for r in resultados if r.get("variacion") == "ERROR"]
    if errores_list:
        story.append(Paragraph(f"Errores de consulta ({len(errores_list)})", sec_s))
        for r in errores_list:
            story.append(Paragraph(
                f"• {r.get('cuit','')} — {r.get('error','')[:80]}",
                ParagraphStyle("E", parent=styles["Normal"],
                               fontSize=8, textColor=colors.HexColor("#cc6600"),
                               leftIndent=10),
            ))

    doc.build(story, onFirstPage=_pie, onLaterPages=_pie)
    return buffer.getvalue()
